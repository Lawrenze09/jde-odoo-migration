"""
reports/migration_report.py

Role in pipeline: Final stage output — generates an Excel reconciliation
report that a non-technical business owner can read and trust. Four sheets:
  Sheet 1 — Summary:           Counts, run status, run mode, timestamp
  Sheet 2 — Validated Records: Every record that passed all 8 business rules
  Sheet 3 — Failed Validation: Every rejected record with plain English reason
  Sheet 4 — Odoo Load Results: What actually happened in Odoo per valid record
                                (LOADED, SKIPPED, FAILED, NOT_PROCESSED, DRY RUN)

Sheet 2 and Sheet 4 answer different questions intentionally:
  Sheet 2 — Did the record pass our data quality rules?
  Sheet 4 — Did the record actually make it into Odoo?

Run Status values shown in Summary:
  SUCCESS          — records created successfully, no failures
  SUCCESS_WITH_SKIPS — some created, some already existed
  NO_OP            — nothing to process, all records already migrated
  FAILED           — batch stopped due to Odoo rejection
  DRY RUN          — no Odoo connection attempted
  PARTIAL          — batch interrupted, some records not processed

Input:  valid_records and failed_records from CustomerValidator
        load_result from OdooLoader (None for dry runs)
Output: output/migration_report_YYYYMMDD_HHMMSS.xlsx

Uses openpyxl for Excel formatting — column widths, bold headers,
color coding, and Text number format to prevent scientific notation.
"""

import os
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from utils.logger import get_logger

logger = get_logger(__name__)

# ── Color constants ─────────────────────────────────────────────────────────
# Header background colors — dark enough for white text to be readable
HEADER_GREEN  = "1F7A4A"   # Sheet 2 — Validated Records header
HEADER_RED    = "C0392B"   # Sheet 3 — Failed Validation header
HEADER_BLUE   = "1A3A5C"   # Sheet 1 — Summary title bar
HEADER_PURPLE = "6C3483"   # Sheet 4 — Odoo Load Results header

# Data row fill colors — light enough for dark text to be readable
ROW_GREEN  = "EAF7EE"   # Light green  — valid/loaded records
ROW_RED    = "FDEDEC"   # Light red    — failed/rejected records
ROW_AMBER  = "FEF9E7"   # Light amber  — skipped records (already in Odoo)
ROW_BLUE   = "EBF5FB"   # Light blue   — not processed (batch stopped early)
ROW_ALT    = "F7F9FC"   # Light gray   — alternating row fill for readability


def _status_reason(status: str, odoo_id: int | None) -> str:
    """
    Generate a plain English reason string for a load status.
    Used in Sheet 4 when no specific Odoo error message is available.

    Args:
        status  (str):       LoadStatus value e.g. 'LOADED', 'SKIPPED'
        odoo_id (int|None):  Odoo partner ID if the record was loaded

    Returns:
        str: Human-readable reason for the Odoo Load Results sheet.
    """
    reasons = {
        "LOADED":        f"Successfully created in Odoo (ID={odoo_id})",
        "SKIPPED":       "Already exists in Odoo — not re-created",
        "FAILED":        "Odoo rejected this record — see error for details",
        "NOT_PROCESSED": "Batch stopped before this record was attempted",
        "DRY RUN":       "Dry run — no data written to Odoo",
    }
    return reasons.get(status, "Unknown status")


# ── Column definitions ──────────────────────────────────────────────────────
# Each tuple: (Excel header label, dict key, column width in characters)

# Sheet 2 — Validated Records: full field set from CustomerTransformer
VALID_COLUMNS = [
    ("JDE AN8",       "_jde_an8",      12),
    ("Name",          "name",          30),
    ("Phone",         "phone",         18),
    ("Street",        "street",        35),
    ("Street 2",      "street2",       25),
    ("City",          "city",          20),
    ("Zip",           "zip",           10),
    ("State Code",    "state_code",    14),
    ("Country Code",  "country_code",  14),
    ("VAT / TIN",     "vat",           20),
    ("Customer Rank", "customer_rank", 14),
    ("Is Company",    "is_company",    12),
    ("Parent AN8",    "parent_an8",    12),
    ("Comment",       "comment",       50),
]

# Sheet 3 — Failed Validation: record data + which rule failed and why
FAILED_COLUMNS = [
    ("JDE AN8",        "_jde_an8",         12),
    ("Name",           "name",             30),
    ("Phone",          "phone",            18),
    ("Street",         "street",           35),
    ("City",           "city",             20),
    ("VAT / TIN",      "vat",              20),
    ("Failed Rule",    "_failed_rule",     28),
    ("Failure Reason", "_failure_reason",  60),
]

# Sheet 4 — Odoo Load Results: what actually happened per record in Odoo
ODOO_LOAD_COLUMNS = [
    ("JDE AN8",     "_jde_an8",      12),
    ("Name",        "name",          30),
    ("Odoo Status", "_odoo_status",  16),
    ("Odoo ID",     "_odoo_id",      12),
    ("Reason",      "_odoo_reason",  55),
]

# Columns that must be formatted as Text to prevent Excel scientific notation.
# Excel auto-converts long numeric strings like 123456789000 → 1.23E+11
# and strips leading zeros from phone numbers like 09171234567 → 9171234567.
TEXT_FORMAT_COLUMNS = {"phone", "vat", "zip", "_jde_an8", "parent_an8", "_odoo_id"}


class MigrationReport:
    """
    Generates a formatted four-sheet Excel reconciliation report.
    One report per pipeline run — timestamped so runs are never overwritten.

    Sheet 1 — Summary:           High-level counts and run metadata
    Sheet 2 — Validated Records: Records that passed all 8 business rules
    Sheet 3 — Failed Validation: Records rejected before Odoo was contacted
    Sheet 4 — Odoo Load Results: What Odoo actually did with each valid record
    """

    def __init__(self, output_dir: str = "output"):
        """
        Initialize the report generator.

        Args:
            output_dir (str): Directory to write report files.
                              Created automatically if it does not exist.
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate(
        self,
        valid_records: list[dict],
        failed_records: list[dict],
        dry_run: bool = True,
        source: str = "mock",
        load_result=None,
    ) -> str:
        """
        Generate the full four-sheet Excel migration report.

        Args:
            valid_records  (list[dict]): Valid records from CustomerValidator.
            failed_records (list[dict]): Failed records from CustomerValidator.
            dry_run        (bool):       True if this was a dry run — no Odoo writes.
            source         (str):        Data source label e.g. 'mock' or 'oracle'.
            load_result:                 LoadResult from OdooLoader, None for dry runs.

        Returns:
            str: Full path to the generated .xlsx file.
        """
        wb = Workbook()

        # openpyxl always creates one default sheet — rename it to Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_valid   = wb.create_sheet("Validated Records")
        ws_failed  = wb.create_sheet("Failed Validation")
        ws_odoo    = wb.create_sheet("Odoo Load Results")

        self._build_summary_sheet(
            ws_summary, valid_records, failed_records,
            dry_run, source, load_result
        )
        self._build_valid_sheet(ws_valid, valid_records)
        self._build_failed_sheet(ws_failed, failed_records)
        self._build_odoo_load_sheet(ws_odoo, valid_records, load_result, dry_run)

        output_path = self._build_output_path()
        wb.save(output_path)

        logger.info(
            f"Migration report generated | "
            f"valid: {len(valid_records)} | "
            f"failed: {len(failed_records)} | "
            f"path: {output_path}"
        )
        return output_path

    # ── Sheet builders ──────────────────────────────────────────────────────

    def _classify_status(
        self,
        load_result,
        dry_run: bool,
    ) -> tuple[str, str]:
        """
        Classify the overall run outcome and return a label and plain English message.
        Used in the Summary sheet Run Status row instead of a raw percentage.

        Status labels:
            DRY_RUN          — no Odoo connection attempted
            NO_OP            — nothing to process, all already migrated
            SUCCESS          — all valid records created, no failures
            SUCCESS_WITH_SKIPS — some created, some already existed
            FAILED           — batch stopped due to Odoo rejection
            PARTIAL          — some records not processed (batch interrupted)

        Args:
            load_result: LoadResult from OdooLoader, or None for dry runs.
            dry_run:     bool — True if no Odoo writes were attempted.

        Returns:
            tuple: (status_label, human_readable_message)
        """
        if dry_run or load_result is None:
            return "DRY_RUN", "Dry run — no data written to Odoo"

        if load_result.total == 0:
            return "NO_OP", "No new records to process — sync is up to date"

        if load_result.failed > 0:
            return (
                "FAILED",
                f"Migration stopped — {load_result.failed} record(s) rejected by Odoo. "
                f"Fix and re-run."
            )

        if load_result.not_processed > 0:
            return (
                "PARTIAL",
                f"Partial migration — {load_result.not_processed} record(s) not processed "
                f"because batch stopped early."
            )

        if load_result.loaded > 0 and load_result.skipped > 0:
            return (
                "SUCCESS_WITH_SKIPS",
                f"Migration complete — {load_result.loaded} created, "
                f"{load_result.skipped} already existed in Odoo"
            )

        if load_result.loaded > 0:
            return (
                "SUCCESS",
                f"Migration complete — {load_result.loaded} records created in Odoo"
            )

        if load_result.loaded == 0 and load_result.skipped > 0:
            return (
                "NO_OP",
                "All records already migrated — nothing new to process"
            )

        return "NO_OP", "No records were processed"

    def _build_summary_sheet(self, ws, valid, failed, dry_run, source, load_result):
        """
        Build Sheet 1 — Summary with run metadata, validation counts,
        Odoo load counts (live runs only), and plain English run status.

        Args:
            ws:          openpyxl Worksheet object
            valid:       list of valid records from CustomerValidator
            failed:      list of failed records from CustomerValidator
            dry_run:     bool — True if no data was written to Odoo
            source:      str — data source label shown in report
            load_result: LoadResult from OdooLoader, or None for dry runs
        """
        run_mode = "DRY RUN — No data written to Odoo" if dry_run else "LIVE RUN — Data written to Odoo"
        run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        total    = len(valid) + len(failed)

        # ── Title bar ──
        ws.merge_cells("A1:C1")
        title_cell           = ws["A1"]
        title_cell.value     = "JDE to Odoo Migration Report"
        title_cell.font      = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill      = PatternFill("solid", fgColor=HEADER_BLUE)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # ── Load counts for summary rows ──
        if load_result and not dry_run:
            loaded_count      = load_result.loaded
            odoo_failed_count = load_result.failed
            not_processed     = load_result.not_processed
            skipped_count     = load_result.skipped
        else:
            loaded_count      = None
            odoo_failed_count = None
            not_processed     = None
            skipped_count     = None

        # ── Classify overall run status ──
        status_label, status_message = self._classify_status(load_result, dry_run)

        # ── Summary row data ──
        summary_rows = [
            ("Run Mode",        run_mode),
            ("Data Source",     source.upper()),
            ("Run Timestamp",   run_time),
            ("",                ""),
            ("Total Extracted", total),
            ("",                ""),
            ("— Validation —",  ""),
            ("Valid Records",   len(valid)),
            ("Failed Records",  len(failed)),
            ("",                ""),
        ]

        # Odoo load section only appears for live runs with a load result
        if load_result and not dry_run:
            summary_rows += [
                ("— Odoo Load —",          ""),
                ("Created in Odoo",        loaded_count),
                ("Rejected by Odoo",       odoo_failed_count),
                ("Not Processed",          not_processed),
                ("Skipped (prev. loaded)", skipped_count),
                ("",                       ""),
            ]

        # Run Status always appears last — plain English classification
        summary_rows.append(("Run Status", status_message))

        # Status label → fill and font color for Run Status row
        STATUS_STYLE = {
            "SUCCESS":          ("EAF7EE", "1F7A4A"),
            "SUCCESS_WITH_SKIPS": ("EAF7EE", "1F7A4A"),
            "NO_OP":            ("EBF5FB", "1A3A5C"),
            "FAILED":           ("FDEDEC", "C0392B"),
            "DRY_RUN":          ("FEF9E7", "B7770D"),
            "PARTIAL":          ("FEF9E7", "B7770D"),
        }

        for i, (label, value) in enumerate(summary_rows, start=2):
            label_cell      = ws.cell(row=i, column=1, value=label)
            value_cell      = ws.cell(row=i, column=2, value=value)
            label_cell.font = Font(bold=True)

            if label == "Valid Records":
                value_cell.fill = PatternFill("solid", fgColor=ROW_GREEN)
                value_cell.font = Font(bold=True, color="1F7A4A")
            elif label == "Failed Records" and len(failed) > 0:
                value_cell.fill = PatternFill("solid", fgColor=ROW_RED)
                value_cell.font = Font(bold=True, color="C0392B")
            elif label == "Created in Odoo" and loaded_count is not None:
                value_cell.fill = PatternFill("solid", fgColor=ROW_GREEN)
                value_cell.font = Font(bold=True, color="1F7A4A")
            elif label == "Rejected by Odoo" and odoo_failed_count:
                value_cell.fill = PatternFill("solid", fgColor=ROW_RED)
                value_cell.font = Font(bold=True, color="C0392B")
            elif label == "Run Status":
                bg, fg = STATUS_STYLE.get(status_label, ("F7F9FC", "2C2C2A"))
                value_cell.fill = PatternFill("solid", fgColor=bg)
                value_cell.font = Font(bold=True, size=11, color=fg)
            elif label == "Run Mode":
                color = "FEF9E7" if dry_run else "EAF7EE"
                value_cell.fill = PatternFill("solid", fgColor=color)
                value_cell.font = Font(
                    bold=True,
                    color="B7770D" if dry_run else "1F7A4A"
                )

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 55

    def _build_valid_sheet(self, ws, valid_records):
        """
        Build Sheet 2 — Validated Records.
        Shows every record that passed all 8 business rules, with all
        transformed field values. Does NOT show Odoo load outcome —
        that is Sheet 4's responsibility.

        Args:
            ws:            openpyxl Worksheet object
            valid_records: list of valid transformed records
        """
        self._write_sheet(
            ws,
            records=valid_records,
            columns=VALID_COLUMNS,
            header_color=HEADER_GREEN,
            row_fill=ROW_GREEN,
        )

    def _build_failed_sheet(self, ws, failed_records):
        """
        Build Sheet 3 — Failed Validation.
        Shows every record rejected by the validator, with the specific
        rule that failed and a plain English explanation a non-technical
        business owner can read and act on.

        Args:
            ws:             openpyxl Worksheet object
            failed_records: list of failed records with _failed_rule
                            and _failure_reason keys from CustomerValidator
        """
        self._write_sheet(
            ws,
            records=failed_records,
            columns=FAILED_COLUMNS,
            header_color=HEADER_RED,
            row_fill=ROW_RED,
        )

    def _build_odoo_load_sheet(self, ws, valid_records, load_result, dry_run: bool):
        """
        Build Sheet 4 — Odoo Load Results.
        Shows what actually happened in Odoo for each validated record.
        This sheet answers the question Sheet 2 cannot: of the records
        that passed validation, which ones actually made it into Odoo?

        Row color reflects the Odoo outcome:
            Green  — LOADED: record was successfully created in Odoo
            Amber  — SKIPPED: record already existed in Odoo (idempotent)
            Red    — FAILED: Odoo rejected the record despite passing validation
            Blue   — NOT_PROCESSED: batch stopped before this record was reached
            Gray   — DRY RUN: no Odoo connection was attempted

        Args:
            ws:            openpyxl Worksheet object
            valid_records: Valid records from CustomerValidator
            load_result:   LoadResult from OdooLoader, or None for dry runs
            dry_run:       True if this was a dry run — no Odoo writes
        """
        # ── Header row ──
        for col_idx, (header, _, width) in enumerate(ODOO_LOAD_COLUMNS, start=1):
            cell           = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor=HEADER_PURPLE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 22

        # Build AN8 → RecordResult lookup so we can match load outcomes
        # to valid records by JDE Address Number
        outcome_by_an8 = {}
        if load_result and load_result.records:
            outcome_by_an8 = {r.an8: r for r in load_result.records}

        # Always load historical IDs — needed for skipped records regardless
        # of whether a load_result exists. Skipped records have odoo_id=None
        # in load_result because the loader doesn't re-query Odoo for them.
        historical_ids = self._get_loaded_odoo_ids()

        # Status → row fill color mapping
        STATUS_COLORS = {
            "LOADED":        ROW_GREEN,   # Successfully created
            "SKIPPED":       ROW_AMBER,   # Already existed — idempotent skip
            "FAILED":        ROW_RED,     # Odoo rejected despite valid data
            "NOT_PROCESSED": ROW_BLUE,    # Batch stopped before this record
            "DRY RUN":       ROW_ALT,     # No Odoo connection attempted
        }

        for row_idx, record in enumerate(valid_records, start=2):
            an8 = record.get("_jde_an8")

            if dry_run or load_result is None:
                # Dry run — no Odoo calls were made
                status  = "DRY RUN"
                odoo_id = None
                reason  = "Dry run — no data written to Odoo"
            else:
                outcome = outcome_by_an8.get(an8)
                if outcome is None:
                    status  = "UNKNOWN"
                    odoo_id = historical_ids.get(an8)
                    reason  = "No load result found for this record"
                else:
                    status  = outcome.status.value
                    # For skipped records, odoo_id may be None in load_result —
                    # fall back to the historical transaction log ID
                    odoo_id = outcome.odoo_id or historical_ids.get(an8)
                    # Use Odoo's error message if available, otherwise
                    # generate a plain English reason from the status
                    reason  = outcome.error or _status_reason(status, odoo_id)

            fill_color = STATUS_COLORS.get(status, ROW_ALT)

            row_data = {
                "_jde_an8":     an8,
                "name":         record.get("name"),
                "_odoo_status": status,
                "_odoo_id":     odoo_id,
                "_odoo_reason": reason,
            }

            for col_idx, (_, key, _) in enumerate(ODOO_LOAD_COLUMNS, start=1):
                value     = row_data.get(key)
                cell      = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = PatternFill("solid", fgColor=fill_color)

                # Force Text format to prevent scientific notation on
                # long numeric IDs like Odoo partner IDs and JDE AN8
                if key in TEXT_FORMAT_COLUMNS:
                    cell.number_format = "@"
                    if value is not None:
                        cell.value = str(value)

        ws.freeze_panes = "A2"

    # ── Shared sheet writer ─────────────────────────────────────────────────

    def _write_sheet(self, ws, records, columns, header_color, row_fill):
        """
        Write a formatted data sheet with bold colored headers and data rows.
        Used by Sheet 2 (Validated Records) and Sheet 3 (Failed Validation).
        Sheet 4 has its own builder because it requires status-based row colors.

        Applies Text number format to phone, vat, zip, and AN8 columns to
        prevent Excel from converting long numbers to scientific notation.

        Args:
            ws:           openpyxl Worksheet
            records:      list[dict] to write
            columns:      list of (header_label, dict_key, column_width) tuples
            header_color: hex color string for header row background
            row_fill:     hex color string for even data rows
        """
        for col_idx, (header, _, width) in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor=header_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 22

        for row_idx, record in enumerate(records, start=2):
            # Alternate row fills for readability when scrolling long lists
            fill_color = row_fill if row_idx % 2 == 0 else ROW_ALT
            for col_idx, (_, key, _) in enumerate(columns, start=1):
                value      = record.get(key)
                cell       = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill  = PatternFill("solid", fgColor=fill_color)

                # Force Text format — prevents scientific notation and
                # preserves leading zeros on phone numbers and TIN fields.
                # Must also re-set the value as str — openpyxl can write
                # an integer to a Text-formatted cell and Excel still converts it.
                if key in TEXT_FORMAT_COLUMNS:
                    cell.number_format = "@"
                    if value is not None:
                        cell.value = str(value)

        ws.freeze_panes = "A2"

    # ── Helpers ─────────────────────────────────────────────────────────────

    def _build_output_path(self) -> str:
        """
        Build a timestamped output file path for the Excel report.
        Timestamp ensures each run produces a unique file — no overwrites.

        Returns:
            str: Full file path e.g. output/migration_report_20260313_201500.xlsx
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename  = f"migration_report_{timestamp}.xlsx"
        return os.path.join(self.output_dir, filename)

    def _get_loaded_odoo_ids(self, db_path: str = "logs/transaction_log.db") -> dict:
        """
        Query the transaction log for Odoo IDs of previously loaded records.
        Used to populate Sheet 4 with Odoo IDs for records skipped this run
        because they were already loaded in a previous run.

        Args:
            db_path (str): Path to the SQLite transaction log.

        Returns:
            dict: AN8 → Odoo partner ID for all LOADED records.
        """
        try:
            with sqlite3.connect(db_path) as conn:
                rows = conn.execute(
                    """
                    SELECT an8, odoo_id FROM migration_log
                    WHERE status = 'LOADED' AND odoo_id IS NOT NULL
                    ORDER BY id DESC
                    """
                ).fetchall()
            # Keep only the most recent Odoo ID per AN8 —
            # dict comprehension with ORDER BY id DESC means first row per AN8 wins
            return {row[0]: row[1] for row in rows}
        except Exception:
            return {}
        